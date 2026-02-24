import io
import json
import os
import re
import uuid as uuid_lib
from datetime import date, datetime

from fastapi import APIRouter, Request, Form, Depends, UploadFile, File
from fastapi.responses import RedirectResponse, HTMLResponse, StreamingResponse, Response
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from ..database import get_db
from ..models import Clan, Clanarina, Aktivnost, Nastavitev
from ..auth import require_login, is_admin, is_editor
from ..config import get_nastavitev, get_tipi_clanstva, get_operaterski_razredi
from ..csrf import get_csrf_token, csrf_protect
from ..audit_log import log_akcija

router = APIRouter(prefix="/izvoz")
templates = Jinja2Templates(directory="app/templates")
templates.env.globals["csrf_token"] = get_csrf_token

MAX_UPLOAD_BYTES = 10 * 1024 * 1024  # 10 MB
DOVOLJENE_PRIPONE = {".xlsx"}
TMP_DIR = "data/tmp"

# ---------------------------------------------------------------------------
# ZRS izvoz – konfigurabilni stolpci, mapiranja, transformacije
# ---------------------------------------------------------------------------

ZRS_CONFIG_KEY = "zrs_izvoz_config"

# Definicija stolpcev ZRS izvoza (fiksni kljuci, nastavljiva imena/vrstni red/vkljuci)
ZRS_STOLPCI_DEF: list[dict] = [
    {"kljuc": "es",           "privzeti_naziv": "ES",                                  "privzeto_vkljuci": True},
    {"kljuc": "priimek",      "privzeti_naziv": "Priimek",                             "privzeto_vkljuci": True},
    {"kljuc": "ime",          "privzeti_naziv": "Ime",                                 "privzeto_vkljuci": True},
    {"kljuc": "klicni_znak",  "privzeti_naziv": "Klicni znak",                        "privzeto_vkljuci": True},
    {"kljuc": "naslov_ulica", "privzeti_naziv": "Naslov - ulica/naselje",              "privzeto_vkljuci": True},
    {"kljuc": "naslov_posta", "privzeti_naziv": "Naslov - pošta",                     "privzeto_vkljuci": True},
    {"kljuc": "tip_clanstva", "privzeti_naziv": "Tip članstva",                       "privzeto_vkljuci": True},
    {"kljuc": "nosilci",      "privzeti_naziv": "Klicni znak nosilci druž. članstva", "privzeto_vkljuci": True},
    {"kljuc": "op_razred",    "privzeti_naziv": "Operaterski razred",                 "privzeto_vkljuci": True},
]

# Privzeto mapiranje tipov članstva: interni naziv → ZRS naziv (prazno = izključi iz izvoza)
ZRS_TIPI_PRIVZETO: dict[str, str] = {
    "Osebni":       "Osebni",
    "Družinski":    "Družinski",
    "Mladi":        "Mladi-18 let",
    "Invalid":      "Invalid",
    "Simpatizerji": "",   # privzeto izključeni
}

_ZRS_COLUMN_WIDTHS: dict[str, int] = {
    "es": 8, "priimek": 18, "ime": 15, "klicni_znak": 12,
    "naslov_ulica": 30, "naslov_posta": 20, "tip_clanstva": 20,
    "nosilci": 22, "op_razred": 16,
}


def _zrs_config_privzeto(db: Session) -> dict:
    """Vrne privzeto ZRS konfiguracijo glede na trenutne nastavitve."""
    tipi = get_tipi_clanstva(db)
    razredi = get_operaterski_razredi(db)
    return {
        "uppercase": True,
        "stolpci": [
            {"kljuc": d["kljuc"], "naziv": d["privzeti_naziv"],
             "vrstni_red": i + 1, "vkljuci": d["privzeto_vkljuci"]}
            for i, d in enumerate(ZRS_STOLPCI_DEF)
        ],
        "tipi": {t: ZRS_TIPI_PRIVZETO.get(t, t) for t in tipi},
        "razredi": {r: r for r in razredi},
    }


def _get_zrs_config(db: Session) -> dict:
    """Prebere ZRS konfiguracijo iz baze (JSON). Ob manjkajočem ali neveljavnem zapisu vrne privzeto."""
    raw = get_nastavitev(db, ZRS_CONFIG_KEY, "")
    if raw:
        try:
            return json.loads(raw)
        except Exception:
            pass
    return _zrs_config_privzeto(db)


def _zrs_vrednost(clan: Clan, kljuc: str, zap_st: int, config: dict):
    """Vrne vrednost stolpca za danega člana z aplikacijo mappingov in uppercase transformacije."""
    tipi: dict = config.get("tipi", {})
    razredi: dict = config.get("razredi", {})
    uppercase: bool = config.get("uppercase", False)

    if kljuc == "es":
        return clan.es_stevilka if clan.es_stevilka else zap_st  # integer – brez uppercase

    val_map: dict[str, str] = {
        "priimek":      clan.priimek or "",
        "ime":          clan.ime or "",
        "klicni_znak":  clan.klicni_znak or "",
        "naslov_ulica": clan.naslov_ulica or "",
        "naslov_posta": clan.naslov_posta or "",
        "tip_clanstva": tipi.get(clan.tip_clanstva or "", clan.tip_clanstva or ""),
        "nosilci":      clan.klicni_znak_nosilci or "",
        "op_razred":    razredi.get(clan.operaterski_razred or "", clan.operaterski_razred or ""),
    }
    val = val_map.get(kljuc, "")
    return val.upper() if uppercase and val else val


# ---------------------------------------------------------------------------
# Uvoz iz Excel – privzeta mapiranja stolpcev Excel → polja baze
# ---------------------------------------------------------------------------

# Privzeta mapiranja stolpcev Excel → polja baze (vejica = alternativna imena)
UVOZ_STOLPCI_PRIVZETO: dict[str, str] = {
    "priimek":            "priimek",
    "ime":                "ime",
    "klicni_znak":        "klicni znak",
    "tip_clanstva":       "tip članstva, tip clanstva, tip",
    "veljavnost_rd":      "veljavnost",
    "naslov_ulica":       "naslov - ulica/naselje, ulica",
    "naslov_posta":       "naslov - pošta, naslov - posta, pošta",
    "nosilci":            "klicni znak nosilci, nosilci",
    "operaterski_razred": "operaterski razred, operaterski",
    "mobilni":            "mobilni telefon, mobilni",
    "telefon_doma":       "telefon doma",
    "email":              "elektronska posta, e-mail, email",
    "soglasje":           "soglasje op, soglasje",
    "izjava":             "izjava",
}


def _get_uvoz_mapping(db: Session) -> dict[str, list[str]]:
    """Prebere mapiranje stolpcev iz nastavitev baze. Vrednosti ločene z vejico."""
    result: dict[str, list[str]] = {}
    for kljuc, privzeto in UVOZ_STOLPCI_PRIVZETO.items():
        vrednost = get_nastavitev(db, f"uvoz_stolpec_{kljuc}", privzeto)
        result[kljuc] = [s.strip() for s in vrednost.split(",") if s.strip()]
    return result


# ---------------------------------------------------------------------------
# Uvoz plačil – privzeta mapiranja stolpcev Excel → polja
# ---------------------------------------------------------------------------

UVOZ_PLACILA_STOLPCI_PRIVZETO: dict[str, str] = {
    "priimek": "priimek",
    "ime":     "ime",
    "datum":   "datum plačila, datum",
    "znesek":  "znesek",
}

KLJUCI_UVOZ_PLACILA = [
    ("uvoz_placila_priimek", "Priimek"),
    ("uvoz_placila_ime",     "Ime"),
    ("uvoz_placila_datum",   "Datum plačila"),
    ("uvoz_placila_znesek",  "Znesek"),
]


def _get_uvoz_placila_mapping(db: Session) -> dict[str, list[str]]:
    """Prebere mapiranje stolpcev plačil iz nastavitev baze."""
    result: dict[str, list[str]] = {}
    for field, privzeto in UVOZ_PLACILA_STOLPCI_PRIVZETO.items():
        vrednost = get_nastavitev(db, f"uvoz_placila_{field}", privzeto)
        result[field] = [s.strip() for s in vrednost.split(",") if s.strip()]
    return result


def _parse_datum_celice(v) -> date | None:
    """Razčleni vrednost Excel celice v datum (datetime, date ali niz)."""
    if v is None:
        return None
    if hasattr(v, "date"):
        return v.date()
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%d. %m. %Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def _parse_excel_placila_pregled(
    vsebina: bytes, db: Session, mapping: dict[str, list[str]]
) -> tuple[list[dict], list[dict]]:
    """Parsira Excel plačil in vrne (za_uvoz, preskoceni) za predogled."""
    ws, headers = _load_sheet(vsebina)
    za_uvoz:    list[dict] = []
    preskoceni: list[dict] = []

    for row in ws.iter_rows(min_row=2):
        priimek = _col(row, headers, *mapping["priimek"])
        ime     = _col(row, headers, *mapping["ime"])
        if not priimek and not ime:
            continue

        priimek_n = priimek.strip().title()
        ime_n     = ime.strip().title()

        # Datum iz stolpca
        datum_val: date | None = None
        for i, h in enumerate(headers):
            if h and any(t.lower() in str(h).strip().lower() for t in mapping["datum"]):
                datum_val = _parse_datum_celice(row[i].value)
                break

        if not datum_val:
            preskoceni.append({
                "priimek": priimek_n, "ime": ime_n,
                "razlog": "Datum ni veljavno",
            })
            continue

        clan = db.query(Clan).filter(
            Clan.priimek == priimek_n, Clan.ime == ime_n
        ).first()
        if not clan:
            preskoceni.append({
                "priimek": priimek_n, "ime": ime_n,
                "razlog": "Član ni najden",
            })
            continue

        leto = datum_val.year
        znesek = _col(row, headers, *mapping["znesek"])
        obstaja = db.query(Clanarina).filter(
            Clanarina.clan_id == clan.id, Clanarina.leto == leto
        ).first()

        za_uvoz.append({
            "priimek":      priimek_n,
            "ime":          ime_n,
            "leto":         leto,
            "datum_placila": datum_val,
            "znesek":       znesek,
            "posodobi":     obstaja is not None,
        })

    return za_uvoz, preskoceni


def _uvozi_placila_workbook(
    vsebina: bytes, db: Session, mapping: dict[str, list[str]]
) -> tuple[int, int]:
    """Uvozi plačila iz Excel. Vrne (uvozeni, preskoceni)."""
    ws, headers = _load_sheet(vsebina)
    uvozeni  = 0
    preskoceni = 0

    for row in ws.iter_rows(min_row=2):
        priimek = _col(row, headers, *mapping["priimek"])
        ime     = _col(row, headers, *mapping["ime"])
        if not priimek and not ime:
            continue

        priimek_n = priimek.strip().title()
        ime_n     = ime.strip().title()

        datum_val: date | None = None
        for i, h in enumerate(headers):
            if h and any(t.lower() in str(h).strip().lower() for t in mapping["datum"]):
                datum_val = _parse_datum_celice(row[i].value)
                break

        if not datum_val:
            preskoceni += 1
            continue

        clan = db.query(Clan).filter(
            Clan.priimek == priimek_n, Clan.ime == ime_n
        ).first()
        if not clan:
            preskoceni += 1
            continue

        leto   = datum_val.year
        znesek = _col(row, headers, *mapping["znesek"]) or None

        obstoječe = db.query(Clanarina).filter(
            Clanarina.clan_id == clan.id, Clanarina.leto == leto
        ).first()
        if obstoječe:
            obstoječe.datum_placila = datum_val
            obstoječe.znesek        = znesek
        else:
            db.add(Clanarina(
                clan_id=clan.id, leto=leto,
                datum_placila=datum_val, znesek=znesek,
            ))
        uvozeni += 1

    db.commit()
    return uvozeni, preskoceni


# Labele za UI (kljuc brez "uvoz_stolpec_" predpone)
KLJUCI_UVOZ = [
    ("uvoz_stolpec_priimek",            "Priimek"),
    ("uvoz_stolpec_ime",                "Ime"),
    ("uvoz_stolpec_klicni_znak",        "Klicni znak"),
    ("uvoz_stolpec_tip_clanstva",       "Tip članstva"),
    ("uvoz_stolpec_veljavnost_rd",      "Veljavnost RD"),
    ("uvoz_stolpec_naslov_ulica",       "Naslov – ulica"),
    ("uvoz_stolpec_naslov_posta",       "Naslov – pošta"),
    ("uvoz_stolpec_nosilci",            "Klicni znak nosilci"),
    ("uvoz_stolpec_operaterski_razred", "Operaterski razred"),
    ("uvoz_stolpec_mobilni",            "Mobilni telefon"),
    ("uvoz_stolpec_telefon_doma",       "Telefon doma"),
    ("uvoz_stolpec_email",              "E-poštni naslov"),
    ("uvoz_stolpec_soglasje",           "Soglasje OP"),
    ("uvoz_stolpec_izjava",             "Izjava"),
]


# ---------------------------------------------------------------------------
# Pomožne funkcije za branje Excel stolpcev
# ---------------------------------------------------------------------------

def _col(row, headers, *iskani) -> str:
    """Vrni vrednost prvega stolpca ki se ujema z enim od iskalnih nizov."""
    # 1. pass: natančno ujemanje
    for iskanje in iskani:
        for i, h in enumerate(headers):
            if h and str(h).strip().lower() == iskanje.lower():
                v = row[i].value
                return str(v).strip() if v is not None else ""
    # 2. pass: vsebuje (celo besedo)
    for iskanje in iskani:
        for i, h in enumerate(headers):
            if h:
                h_l = str(h).strip().lower()
                s_l = iskanje.lower()
                if re.search(r'(?<![a-zčšž])' + re.escape(s_l) + r'(?![a-zčšž])', h_l):
                    v = row[i].value
                    return str(v).strip() if v is not None else ""
    return ""


def _parse_rd(row, headers, iskani: list[str]) -> date | None:
    """Vrne datum veljavnosti RD ali None. iskani = seznam možnih imen stolpca."""
    for i, h in enumerate(headers):
        if h and any(t.lower() in str(h).strip().lower() for t in iskani):
            v = row[i].value
            if v:
                if hasattr(v, "date"):
                    return v.date()
                else:
                    try:
                        return datetime.strptime(str(v), "%Y-%m-%d").date()
                    except Exception:
                        pass
            break
    return None


def _parse_datum_placila(row, headers, leto: int) -> date | None:
    """Vrne datum plačila za izbrano leto ali None."""
    for i, h in enumerate(headers):
        if h and str(leto) in str(h):
            v = row[i].value
            if v:
                if hasattr(v, "date"):
                    return v.date()
                elif str(v).strip().lower() not in ("", "ne", "n", "0"):
                    try:
                        return datetime.strptime(str(v), "%Y-%m-%d").date()
                    except Exception:
                        return date(leto, 1, 1)
            break
    return None


def _load_sheet(vsebina: bytes):
    """Naloži Excel in vrne (ws, headers)."""
    wb = load_workbook(io.BytesIO(vsebina), data_only=True, read_only=True)
    list_ime = None
    for ime in wb.sheetnames:
        if "ListaVsi" in ime or "Lista" in ime:
            list_ime = ime
            break
    if not list_ime:
        list_ime = wb.sheetnames[0]
    ws = wb[list_ime]
    headers = [cell.value for cell in ws[1]]
    return ws, headers


def _parse_excel_pregled(vsebina: bytes, db: Session, mapping: dict[str, list[str]]) -> tuple[list[dict], list[dict]]:
    """Parsira Excel in vrne (novi, preskoceni) za predogled, brez pisanja v DB."""
    ws, headers = _load_sheet(vsebina)
    tipi = get_tipi_clanstva(db)
    novi: list[dict] = []
    preskoceni: list[dict] = []

    for row in ws.iter_rows(min_row=2):
        priimek = _col(row, headers, *mapping["priimek"])
        ime = _col(row, headers, *mapping["ime"])
        if not priimek or not ime:
            continue

        priimek_n = priimek.strip().title()
        ime_n = ime.strip().title()

        obstaja = db.query(Clan).filter(
            Clan.priimek == priimek_n, Clan.ime == ime_n
        ).first()

        tip = _col(row, headers, *mapping["tip_clanstva"])
        if tip not in tipi:
            tip = tipi[0] if tipi else "Osebni"

        entry = {
            "priimek": priimek_n,
            "ime": ime_n,
            "klicni_znak": _col(row, headers, *mapping["klicni_znak"]).strip().upper() or "–",
            "tip_clanstva": tip,
            "veljavnost_rd": _parse_rd(row, headers, mapping["veljavnost_rd"]),
        }

        if obstaja:
            preskoceni.append(entry)
        else:
            novi.append(entry)

    return novi, preskoceni


def _uvozi_workbook(vsebina: bytes, db: Session, mapping: dict[str, list[str]]) -> tuple[int, int]:
    """Izvede dejanski uvoz članov iz Excel vsebine. Vrne (uvozeni, preskoceni)."""
    ws, headers = _load_sheet(vsebina)
    tipi = get_tipi_clanstva(db)
    uvozeni = 0
    preskoceni = 0

    for row in ws.iter_rows(min_row=2):
        priimek = _col(row, headers, *mapping["priimek"])
        ime = _col(row, headers, *mapping["ime"])
        if not priimek or not ime:
            continue

        priimek_n = priimek.strip().title()
        ime_n = ime.strip().title()

        obstaja = db.query(Clan).filter(
            Clan.priimek == priimek_n, Clan.ime == ime_n
        ).first()
        if obstaja:
            preskoceni += 1
            continue

        tip = _col(row, headers, *mapping["tip_clanstva"])
        if tip not in tipi:
            tip = tipi[0] if tipi else "Osebni"

        clan = Clan(
            priimek=priimek_n,
            ime=ime_n,
            klicni_znak=(_col(row, headers, *mapping["klicni_znak"]).strip().upper() or None),
            naslov_ulica=_col(row, headers, *mapping["naslov_ulica"]) or None,
            naslov_posta=_col(row, headers, *mapping["naslov_posta"]) or None,
            tip_clanstva=tip,
            klicni_znak_nosilci=_col(row, headers, *mapping["nosilci"]) or None,
            operaterski_razred=_col(row, headers, *mapping["operaterski_razred"]) or None,
            mobilni_telefon=_col(row, headers, *mapping["mobilni"]) or None,
            telefon_doma=_col(row, headers, *mapping["telefon_doma"]) or None,
            elektronska_posta=_col(row, headers, *mapping["email"]) or None,
            soglasje_op=_col(row, headers, *mapping["soglasje"]) or None,
            izjava=_col(row, headers, *mapping["izjava"]) or None,
            veljavnost_rd=_parse_rd(row, headers, mapping["veljavnost_rd"]),
            aktiven=True,
        )
        db.add(clan)
        uvozeni += 1

    db.commit()
    return uvozeni, preskoceni


def _sestevki_placil(db: Session) -> list[dict]:
    """Vrne seštevke plačil po letu: leto, st_placnikov, skupaj_znesek, brez_zneska."""
    from collections import defaultdict
    vsa = db.query(Clanarina).filter(Clanarina.datum_placila != None).all()
    po_letu: dict = defaultdict(lambda: {"st_placnikov": 0, "skupaj": 0.0, "brez_zneska": 0})
    for c in vsa:
        po_letu[c.leto]["st_placnikov"] += 1
        if c.znesek and c.znesek.strip():
            try:
                po_letu[c.leto]["skupaj"] += float(c.znesek.replace(",", "."))
            except ValueError:
                po_letu[c.leto]["brez_zneska"] += 1
        else:
            po_letu[c.leto]["brez_zneska"] += 1
    return [
        {"leto": leto, **vrednosti}
        for leto, vrednosti in sorted(po_letu.items(), reverse=True)
    ]


# ---------------------------------------------------------------------------
# Route handlers
# ---------------------------------------------------------------------------

@router.get("", response_class=HTMLResponse)
async def izvoz_stran(request: Request, db: Session = Depends(get_db)) -> Response:
    user, redirect = require_login(request)
    if redirect:
        return redirect
    leto = date.today().year
    return templates.TemplateResponse(
        "izvoz/index.html",
        {
            "request": request,
            "user": user,
            "leto": leto,
            "sestevki": _sestevki_placil(db),
            "zrs_config": _get_zrs_config(db),
            "zrs_stolpci_def": ZRS_STOLPCI_DEF,
            "zrs_shranjeno": request.query_params.get("zrs_shranjeno") == "1",
            "is_admin": is_admin(user),
            "is_editor": is_editor(user),
        },
    )


@router.get("/zrs")
async def izvoz_zrs(
    request: Request,
    leto: int = date.today().year,
    db: Session = Depends(get_db),
) -> Response:
    user, redirect = require_login(request)
    if redirect:
        return redirect

    config = _get_zrs_config(db)

    # Tipi ki so vključeni v izvoz (neprazen ZRS naziv)
    vkljuceni_tipi = {t for t, zrs_t in config["tipi"].items() if zrs_t}

    placali_ids = {
        c.clan_id
        for c in db.query(Clanarina)
        .filter(Clanarina.leto == leto, Clanarina.datum_placila != None)
        .all()
    }

    if vkljuceni_tipi:
        clani = [
            c for c in db.query(Clan)
            .filter(Clan.aktiven == True, Clan.tip_clanstva.in_(vkljuceni_tipi))
            .order_by(Clan.priimek, Clan.ime)
            .all()
            if c.id in placali_ids
        ]
    else:
        clani = []

    # Stolpci v želenem vrstnem redu (samo vključeni)
    aktivni_stolpci = sorted(
        [s for s in config["stolpci"] if s["vkljuci"]],
        key=lambda s: s["vrstni_red"],
    )

    wb = Workbook()

    # Sheet 1: Podatki radiokluba
    ws_klub = wb.active
    ws_klub.title = "Podatki radiokluba"
    klub_ime = get_nastavitev(db, "klub_ime", "")
    klub_oznaka = get_nastavitev(db, "klub_oznaka", "")
    klub_naslov = get_nastavitev(db, "klub_naslov", "")
    klub_posta = get_nastavitev(db, "klub_posta", "")
    klub_email = get_nastavitev(db, "klub_email", "")
    ws_klub.append(["Ime radiokluba:", klub_ime])
    ws_klub.append(["Klicni znak:", klub_oznaka])
    ws_klub.append(["Naslov:", klub_naslov])
    ws_klub.append(["Pošta:", klub_posta])
    ws_klub.append(["E-pošta:", klub_email])
    ws_klub.append(["Leto prijave:", leto])
    ws_klub.column_dimensions["A"].width = 20
    ws_klub.column_dimensions["B"].width = 40

    # Sheet 2: Lista članov
    ws = wb.create_sheet("ListaClanov")
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDEEFF")
    ws.append([s["naziv"] for s in aktivni_stolpci])
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for zap_st, clan in enumerate(clani, start=1):
        ws.append([_zrs_vrednost(clan, s["kljuc"], zap_st, config) for s in aktivni_stolpci])

    for ci, stolpec in enumerate(aktivni_stolpci, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = _ZRS_COLUMN_WIDTHS.get(stolpec["kljuc"], 15)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"prijava_clanov_{leto}_{klub_oznaka}.xlsx"
    ip = request.client.host if request.client else None
    log_akcija(db, user.get("uporabnisko_ime") if user else None, "izvoz_zrs",
               f"Leto {leto}", ip=ip)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/zrs-nastavitve")
async def zrs_nastavitve_shrani(
    request: Request,
    db: Session = Depends(get_db),
    _csrf: None = Depends(csrf_protect),
) -> Response:
    user, redirect = require_login(request)
    if redirect:
        return redirect
    if not is_admin(user):
        return RedirectResponse(url="/izvoz", status_code=302)

    form = await request.form()

    uppercase = form.get("zrs_uppercase") == "1"

    tipi: dict[str, str] = {}
    i = 0
    while form.get(f"tip_kljuc_{i}") is not None:
        kljuc = str(form.get(f"tip_kljuc_{i}"))
        tipi[kljuc] = str(form.get(f"tip_vrednost_{i}", "")).strip()
        i += 1

    razredi: dict[str, str] = {}
    i = 0
    while form.get(f"razred_kljuc_{i}") is not None:
        kljuc = str(form.get(f"razred_kljuc_{i}"))
        razredi[kljuc] = str(form.get(f"razred_vrednost_{i}", "")).strip()
        i += 1

    stolpci: list[dict] = []
    i = 0
    while form.get(f"stolpec_kljuc_{i}") is not None:
        stolpci.append({
            "kljuc":      str(form.get(f"stolpec_kljuc_{i}")),
            "naziv":      str(form.get(f"stolpec_naziv_{i}", "")).strip(),
            "vrstni_red": int(form.get(f"stolpec_red_{i}", i + 1)),  # type: ignore[arg-type]
            "vkljuci":    form.get(f"stolpec_vkljuci_{i}") == "1",
        })
        i += 1

    config = {"uppercase": uppercase, "tipi": tipi, "razredi": razredi, "stolpci": stolpci}
    raw = json.dumps(config, ensure_ascii=False)

    n = db.query(Nastavitev).filter(Nastavitev.kljuc == ZRS_CONFIG_KEY).first()
    if n:
        n.vrednost = raw
    else:
        db.add(Nastavitev(kljuc=ZRS_CONFIG_KEY, vrednost=raw, opis="Konfiguracija ZRS izvoza (JSON)"))
    db.commit()

    return RedirectResponse(url="/izvoz?zrs_shranjeno=1", status_code=302)


@router.get("/backup-excel")
async def backup_excel(request: Request, db: Session = Depends(get_db)) -> Response:
    user, redirect = require_login(request)
    if redirect:
        return redirect

    clani = db.query(Clan).order_by(Clan.priimek, Clan.ime).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Clani"
    header_font = Font(bold=True)

    # Glava
    headers = [
        "ID", "Priimek", "Ime", "Klicni znak", "Naslov - ulica/naselje",
        "Naslov - pošta", "Tip članstva", "Klicni znak nosilci",
        "Operaterski razred", "Mobilni telefon", "Telefon doma",
        "E-pošta", "Soglasje OP", "Izjava", "Veljavnost RD",
        "ES številka", "Aktiven", "Opombe",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font

    for clan in clani:
        ws.append([
            clan.id,
            clan.priimek,
            clan.ime,
            clan.klicni_znak or "",
            clan.naslov_ulica or "",
            clan.naslov_posta or "",
            clan.tip_clanstva,
            clan.klicni_znak_nosilci or "",
            clan.operaterski_razred or "",
            clan.mobilni_telefon or "",
            clan.telefon_doma or "",
            clan.elektronska_posta or "",
            clan.soglasje_op or "",
            clan.izjava or "",
            clan.veljavnost_rd.isoformat() if clan.veljavnost_rd else "",
            clan.es_stevilka or "",
            "Da" if clan.aktiven else "Ne",
            clan.opombe or "",
        ])

    # Sheet za clanarine
    ws2 = wb.create_sheet("Clanarine")
    ws2.append(["ID", "Clan ID", "Priimek", "Ime", "Leto", "Datum plačila", "Znesek (€)", "Opombe"])
    ws2[1][0].font = header_font
    clanarine = db.query(Clanarina).order_by(Clanarina.clan_id, Clanarina.leto.desc()).all()
    clan_map = {c.id: c for c in clani}
    for c in clanarine:
        clan = clan_map.get(c.clan_id)
        ws2.append([
            c.id,
            c.clan_id,
            clan.priimek if clan else "",
            clan.ime if clan else "",
            c.leto,
            c.datum_placila.isoformat() if c.datum_placila else "",
            c.znesek or "",
            c.opombe or "",
        ])

    # Sheet za aktivnosti
    ws3 = wb.create_sheet("Aktivnosti")
    ws3.append(["ID", "Clan ID", "Priimek", "Ime", "Leto", "Datum aktivnosti", "Aktivnost", "Delovne ure"])
    ws3[1][0].font = header_font
    aktivnosti = db.query(Aktivnost).order_by(
        Aktivnost.clan_id, Aktivnost.leto.desc(), Aktivnost.datum.desc()
    ).all()
    for a in aktivnosti:
        clan = clan_map.get(a.clan_id)
        ws3.append([
            a.id,
            a.clan_id,
            clan.priimek if clan else "",
            clan.ime if clan else "",
            a.leto,
            a.datum.isoformat() if a.datum else "",
            a.opis,
            a.delovne_ure if a.delovne_ure is not None else "",
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    today = date.today().isoformat()
    ip = request.client.host if request.client else None
    log_akcija(db, user.get("uporabnisko_ime") if user else None, "izvoz_backup_excel", ip=ip)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="backup_clanstvo_{today}.xlsx"'},
    )


@router.get("/backup-db")
async def backup_db(request: Request, db: Session = Depends(get_db)) -> Response:
    user, redirect = require_login(request)
    if redirect:
        return redirect
    if not is_admin(user):
        return RedirectResponse(url="/izvoz", status_code=302)

    db_path = "data/clanstvo.db"
    if not os.path.exists(db_path):
        return RedirectResponse(url="/izvoz", status_code=302)

    today = date.today().isoformat()
    ip = request.client.host if request.client else None
    log_akcija(db, user.get("uporabnisko_ime") if user else None, "izvoz_backup_db", ip=ip)

    def iter_file():
        with open(db_path, "rb") as f:
            yield from f

    return StreamingResponse(
        iter_file(),
        media_type="application/octet-stream",
        headers={"Content-Disposition": f'attachment; filename="backup_clanstvo_{today}.db"'},
    )


@router.get("/uvozi", response_class=HTMLResponse)
async def uvozi_stran(request: Request, db: Session = Depends(get_db)) -> Response:
    user, redirect = require_login(request)
    if redirect:
        return redirect
    if not is_admin(user):
        return RedirectResponse(url="/izvoz", status_code=302)

    nas = {n.kljuc: n.vrednost for n in db.query(Nastavitev).filter(
        Nastavitev.kljuc.like("uvoz_stolpec_%")
    ).all()}
    for kljuc, _ in KLJUCI_UVOZ:
        if kljuc not in nas or not nas[kljuc]:
            field = kljuc[len("uvoz_stolpec_"):]
            nas[kljuc] = UVOZ_STOLPCI_PRIVZETO.get(field, "")

    placila_nas = {n.kljuc: n.vrednost for n in db.query(Nastavitev).filter(
        Nastavitev.kljuc.like("uvoz_placila_%")
    ).all()}
    for kljuc, _ in KLJUCI_UVOZ_PLACILA:
        if kljuc not in placila_nas or not placila_nas[kljuc]:
            field = kljuc[len("uvoz_placila_"):]
            placila_nas[kljuc] = UVOZ_PLACILA_STOLPCI_PRIVZETO.get(field, "")

    return templates.TemplateResponse(
        "izvoz/uvoz.html",
        {
            "request": request,
            "user": user,
            "is_admin": True,
            "kljuci_uvoz": KLJUCI_UVOZ,
            "uvoz_nas": nas,
            "uvoz_shranjeno": request.query_params.get("uvoz_shranjeno") == "1",
            "kljuci_uvoz_placila": KLJUCI_UVOZ_PLACILA,
            "placila_nas": placila_nas,
            "placila_shranjeno": request.query_params.get("placila_shranjeno") == "1",
        },
    )


@router.post("/uvozi-nastavitve")
async def uvozi_nastavitve_shrani(
    request: Request,
    db: Session = Depends(get_db),
    _csrf: None = Depends(csrf_protect),
) -> Response:
    user, redirect = require_login(request)
    if redirect:
        return redirect
    if not is_admin(user):
        return RedirectResponse(url="/izvoz", status_code=302)

    form = await request.form()
    for kljuc, _ in KLJUCI_UVOZ:
        vrednost = str(form.get(kljuc, "")).strip()
        n = db.query(Nastavitev).filter(Nastavitev.kljuc == kljuc).first()
        if n:
            n.vrednost = vrednost
        else:
            db.add(Nastavitev(kljuc=kljuc, vrednost=vrednost))
    db.commit()
    return RedirectResponse(url="/izvoz/uvozi?uvoz_shranjeno=1", status_code=302)


@router.post("/uvozi", response_class=HTMLResponse)
async def uvozi_pregled(
    request: Request,
    datoteka: UploadFile = File(...),
    db: Session = Depends(get_db),
    _csrf: None = Depends(csrf_protect),
) -> Response:
    user, redirect = require_login(request)
    if redirect:
        return redirect
    if not is_admin(user):
        return RedirectResponse(url="/izvoz", status_code=302)

    def _uvoz_err(napaka: str):
        return templates.TemplateResponse(
            "izvoz/uvoz.html",
            {"request": request, "user": user, "is_admin": True,
             "napaka": napaka, "kljuci_uvoz": KLJUCI_UVOZ,
             "uvoz_nas": {}, "uvoz_shranjeno": False,
             "kljuci_uvoz_placila": KLJUCI_UVOZ_PLACILA,
             "placila_nas": {}, "placila_shranjeno": False},
        )

    ime = (datoteka.filename or "").lower()
    if not any(ime.endswith(p) for p in DOVOLJENE_PRIPONE):
        return _uvoz_err("Dovoljena je samo datoteka .xlsx.")

    vsebina = await datoteka.read()
    if len(vsebina) > MAX_UPLOAD_BYTES:
        return _uvoz_err("Datoteka je prevelika (max 10 MB).")

    mapping = _get_uvoz_mapping(db)
    novi, preskoceni = _parse_excel_pregled(vsebina, db, mapping)

    os.makedirs(TMP_DIR, exist_ok=True)
    uvoz_uuid = str(uuid_lib.uuid4())
    tmp_pot = os.path.join(TMP_DIR, f"{uvoz_uuid}.xlsx")
    with open(tmp_pot, "wb") as f:
        f.write(vsebina)

    request.session["_uvoz_uuid"] = uvoz_uuid

    return templates.TemplateResponse(
        "izvoz/uvoz-pregled.html",
        {
            "request": request,
            "user": user,
            "is_admin": True,
            "novi": novi,
            "preskoceni": preskoceni,
        },
    )


@router.post("/uvozi-potrdi", response_class=HTMLResponse)
async def uvozi_potrdi(
    request: Request,
    db: Session = Depends(get_db),
    _csrf: None = Depends(csrf_protect),
) -> Response:
    user, redirect = require_login(request)
    if redirect:
        return redirect
    if not is_admin(user):
        return RedirectResponse(url="/izvoz", status_code=302)

    uvoz_uuid = request.session.get("_uvoz_uuid")

    def _potrdi_err(napaka: str):
        return templates.TemplateResponse(
            "izvoz/uvoz.html",
            {"request": request, "user": user, "is_admin": True,
             "napaka": napaka, "kljuci_uvoz": KLJUCI_UVOZ,
             "uvoz_nas": {}, "uvoz_shranjeno": False,
             "kljuci_uvoz_placila": KLJUCI_UVOZ_PLACILA,
             "placila_nas": {}, "placila_shranjeno": False},
        )

    if not uvoz_uuid:
        return _potrdi_err("Seja je potekla. Prosim naložite datoteko znova.")

    tmp_pot = os.path.join(TMP_DIR, f"{uvoz_uuid}.xlsx")
    if not os.path.exists(tmp_pot):
        return _potrdi_err("Začasna datoteka ni najdena. Prosim naložite datoteko znova.")

    uvozeni = 0
    preskoceni_st = 0
    try:
        with open(tmp_pot, "rb") as f:
            vsebina = f.read()
        mapping = _get_uvoz_mapping(db)
        uvozeni, preskoceni_st = _uvozi_workbook(vsebina, db, mapping)
        ip = request.client.host if request.client else None
        log_akcija(
            db, user.get("uporabnisko_ime") if user else None,
            "uvoz_clanov",
            f"{uvozeni} uvoženih, {preskoceni_st} preskočenih",
            ip=ip,
        )
    finally:
        if os.path.exists(tmp_pot):
            os.remove(tmp_pot)
        request.session.pop("_uvoz_uuid", None)

    return templates.TemplateResponse(
        "izvoz/uvoz.html",
        {
            "request": request,
            "user": user,
            "is_admin": True,
            "sporocilo": f"Uvoz zaključen: {uvozeni} uvoženih, {preskoceni_st} preskočenih (duplikati).",
            "kljuci_uvoz": KLJUCI_UVOZ,
            "uvoz_nas": {},
            "uvoz_shranjeno": False,
            "kljuci_uvoz_placila": KLJUCI_UVOZ_PLACILA,
            "placila_nas": {},
            "placila_shranjeno": False,
        },
    )


@router.post("/uvozi-placila", response_class=HTMLResponse)
async def uvozi_placila_pregled(
    request: Request,
    datoteka: UploadFile = File(...),
    db: Session = Depends(get_db),
    _csrf: None = Depends(csrf_protect),
) -> Response:
    user, redirect = require_login(request)
    if redirect:
        return redirect
    if not is_admin(user):
        return RedirectResponse(url="/izvoz", status_code=302)

    def _placila_err(napaka: str):
        return templates.TemplateResponse(
            "izvoz/uvoz.html",
            {"request": request, "user": user, "is_admin": True,
             "napaka_placila": napaka, "kljuci_uvoz": KLJUCI_UVOZ,
             "uvoz_nas": {}, "uvoz_shranjeno": False,
             "kljuci_uvoz_placila": KLJUCI_UVOZ_PLACILA,
             "placila_nas": {}, "placila_shranjeno": False},
        )

    ime = (datoteka.filename or "").lower()
    if not any(ime.endswith(p) for p in DOVOLJENE_PRIPONE):
        return _placila_err("Dovoljena je samo datoteka .xlsx.")

    vsebina = await datoteka.read()
    if len(vsebina) > MAX_UPLOAD_BYTES:
        return _placila_err("Datoteka je prevelika (max 10 MB).")

    mapping = _get_uvoz_placila_mapping(db)
    za_uvoz, preskoceni = _parse_excel_placila_pregled(vsebina, db, mapping)

    os.makedirs(TMP_DIR, exist_ok=True)
    uvoz_uuid = str(uuid_lib.uuid4())
    tmp_pot = os.path.join(TMP_DIR, f"{uvoz_uuid}.xlsx")
    with open(tmp_pot, "wb") as f:
        f.write(vsebina)

    request.session["_placila_uuid"] = uvoz_uuid

    return templates.TemplateResponse(
        "izvoz/uvoz-placila-pregled.html",
        {
            "request": request,
            "user": user,
            "is_admin": True,
            "za_uvoz": za_uvoz,
            "preskoceni": preskoceni,
        },
    )


@router.post("/uvozi-placila-potrdi", response_class=HTMLResponse)
async def uvozi_placila_potrdi(
    request: Request,
    db: Session = Depends(get_db),
    _csrf: None = Depends(csrf_protect),
) -> Response:
    user, redirect = require_login(request)
    if redirect:
        return redirect
    if not is_admin(user):
        return RedirectResponse(url="/izvoz", status_code=302)

    uvoz_uuid = request.session.get("_placila_uuid")

    def _potrdi_err(napaka: str):
        return templates.TemplateResponse(
            "izvoz/uvoz.html",
            {"request": request, "user": user, "is_admin": True,
             "napaka_placila": napaka, "kljuci_uvoz": KLJUCI_UVOZ,
             "uvoz_nas": {}, "uvoz_shranjeno": False,
             "kljuci_uvoz_placila": KLJUCI_UVOZ_PLACILA,
             "placila_nas": {}, "placila_shranjeno": False},
        )

    if not uvoz_uuid:
        return _potrdi_err("Seja je potekla. Prosim naložite datoteko znova.")

    tmp_pot = os.path.join(TMP_DIR, f"{uvoz_uuid}.xlsx")
    if not os.path.exists(tmp_pot):
        return _potrdi_err("Začasna datoteka ni najdena. Prosim naložite datoteko znova.")

    uvozeni = 0
    preskoceni_st = 0
    try:
        with open(tmp_pot, "rb") as f:
            vsebina = f.read()
        mapping = _get_uvoz_placila_mapping(db)
        uvozeni, preskoceni_st = _uvozi_placila_workbook(vsebina, db, mapping)
        ip = request.client.host if request.client else None
        log_akcija(
            db, user.get("uporabnisko_ime") if user else None,
            "uvoz_placil",
            f"{uvozeni} uvoženih, {preskoceni_st} preskočenih",
            ip=ip,
        )
    finally:
        if os.path.exists(tmp_pot):
            os.remove(tmp_pot)
        request.session.pop("_placila_uuid", None)

    return templates.TemplateResponse(
        "izvoz/uvoz.html",
        {
            "request": request,
            "user": user,
            "is_admin": True,
            "sporocilo_placila": f"Uvoz plačil zaključen: {uvozeni} uvoženih, {preskoceni_st} preskočenih.",
            "kljuci_uvoz": KLJUCI_UVOZ,
            "uvoz_nas": {},
            "uvoz_shranjeno": False,
            "kljuci_uvoz_placila": KLJUCI_UVOZ_PLACILA,
            "placila_nas": {},
            "placila_shranjeno": False,
        },
    )


@router.post("/uvozi-placila-nastavitve")
async def uvozi_placila_nastavitve_shrani(
    request: Request,
    db: Session = Depends(get_db),
    _csrf: None = Depends(csrf_protect),
) -> Response:
    user, redirect = require_login(request)
    if redirect:
        return redirect
    if not is_admin(user):
        return RedirectResponse(url="/izvoz", status_code=302)

    form = await request.form()
    for kljuc, _ in KLJUCI_UVOZ_PLACILA:
        vrednost = str(form.get(kljuc, "")).strip()
        n = db.query(Nastavitev).filter(Nastavitev.kljuc == kljuc).first()
        if n:
            n.vrednost = vrednost
        else:
            db.add(Nastavitev(kljuc=kljuc, vrednost=vrednost))
    db.commit()
    return RedirectResponse(url="/izvoz/uvozi?placila_shranjeno=1", status_code=302)
